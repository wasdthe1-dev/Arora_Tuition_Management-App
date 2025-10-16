import csv
import datetime
import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
from app.config import COLORS, FONTS
from app.ui.components import GoldButton, Card, style_treeview, PanelSwitcher


class AdminApp(ctk.CTkFrame):
    def __init__(self, master, db, on_logout, on_theme_change=None):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db
        self.on_logout = on_logout
        self.on_theme_change = on_theme_change

        # Layout
        self.sidebar = ctk.CTkFrame(self, width=220, fg_color=COLORS["panel"], corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.content = PanelSwitcher(self)
        self.content.pack(side="right", fill="both", expand=True)

        title = ctk.CTkLabel(self.sidebar, text="Admin Panel", font=FONTS["h2"], text_color=COLORS["gold"]) 
        title.pack(pady=(12, 6))

        # Theme toggle (Aqua / Light / Dark)
        mode_switch = ctk.CTkSegmentedButton(
            self.sidebar,
            values=["Aqua", "Light", "Dark"],
        )
        from app.config import CURRENT_MODE
        mode_switch.set("Aqua" if CURRENT_MODE=="aqua" else ("Light" if CURRENT_MODE=="light" else "Dark"))
        def _set_mode(val):
            from app.config import set_theme
            theme = val.lower()
            set_theme(theme)
            if self.on_theme_change:
                self.on_theme_change(theme)
            # rebuild views to apply palette
            self._rebuild_views()
        mode_switch.configure(command=_set_mode)
        mode_switch.pack(fill="x", padx=12, pady=(0, 12))

        self._rebuild_views()

        for name in self.views:
            btn = ctk.CTkButton(
                self.sidebar,
                text=name,
                fg_color=COLORS["gold"],
                hover_color=COLORS["gold_dim"],
                text_color=COLORS["bg1"],
                corner_radius=8,
                command=lambda n=name: self.show(n),
            )
            btn.pack(fill="x", padx=12, pady=6)

        # Logout bottom-left
        spacer = ctk.CTkFrame(self.sidebar, fg_color=COLORS["panel"])
        spacer.pack(expand=True, fill="both")
        ctk.CTkButton(
            self.sidebar,
            text="Logout",
            fg_color="#333333",
            hover_color="#444444",
            command=self._logout,
        ).pack(side="left", anchor="sw", padx=12, pady=12)

        self.current_view = None
        self.show("Dashboard", animate=False)

    def _rebuild_views(self):
        # apply new palette to chrome
        from app.config import COLORS
        self.configure(fg_color=COLORS["bg1"])
        self.sidebar.configure(fg_color=COLORS["panel"])
        self.content.configure(fg_color=COLORS["bg1"])
        # recreate view instances to apply new COLORS
        self.views = {
            "Dashboard": DashboardView(self.content, self.db),
            "Students": StudentsView(self.content, self.db),
            "Batches": BatchesView(self.content, self.db),
            "Time Table": TimetableView(self.content, self.db),
            "Attendance": AttendanceView(self.content, self.db),
            "Fees": FeesView(self.content, self.db),
            "Messages": MessagesView(self.content, self.db),
        }

    def show(self, name: str, animate: bool = True):
        view = self.views[name]
        if self.current_view is view:
            # Already on this page: just refresh
            view.refresh()
            return
        view.refresh()
        if self.current_view is None or not animate:
            self.content.set(view)
        else:
            self.content.transition_to(view, direction=1)
        self.current_view = view

    def _logout(self):
        self.on_logout()

    def _notify_timetable_update(self):
        # Send a broadcast notification when timetable is updated
        from datetime import datetime
        self.db.send_message("Timetable updated", datetime.now().isoformat(sep=' ', timespec='seconds'), "admin", "all")


class DashboardView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"])  
        self.db = db
        self.cards = []
        self.bottom_frames = []

    def refresh(self):
        for c in self.cards:
            c.destroy()
        self.cards.clear()
        for f in self.bottom_frames:
            f.destroy()
        self.bottom_frames.clear()
        # Metrics
        total_students = len(self.db.list_students())
        total_batches = len(self.db.list_batches())
        fees_collected = self.db.get_fees() or 0
        attendance_pct = self.db.attendance_percentage()

        grid = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        grid.pack(fill="x", padx=16, pady=(16, 8))
        metrics = [
            ("Total Students", str(total_students)),
            ("Total Batches", str(total_batches)),
            ("Fees Collected", f"₹ {fees_collected:.2f}"),
            ("Attendance %", f"{attendance_pct}%"),
        ]
        for i, (t, v) in enumerate(metrics):
            card = Card(grid, t, v)
            card.grid(row=0, column=i, padx=8, pady=8, sticky="nsew")
            self.cards.append(card)
        for i in range(len(metrics)):
            grid.grid_columnconfigure(i, weight=1)

        # Recent students
        recent_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        recent_frame.pack(fill="both", expand=True, padx=16, pady=8)
        ctk.CTkLabel(recent_frame, text="Recent Students", text_color=COLORS["gold"], font=FONTS["h2"]).pack(anchor="w", padx=12, pady=6)
        cols = ("ID", "Name", "Class", "Batch")
        tv = ttk.Treeview(recent_frame, columns=cols, show="headings", height=6)
        for c in cols:
            tv.heading(c, text=c)
            tv.column(c, width=140, anchor="w")
        tv.pack(fill="both", expand=True, padx=12, pady=8)
        style_treeview(tv)
        rows = self.db.list_students()
        for r in rows[-10:]:
            tv.insert('', 'end', values=(r[0], r[1], r[3], r[7]))
        self.bottom_frames.append(recent_frame)

        # Announcements preview
        ann_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        ann_frame.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        ctk.CTkLabel(ann_frame, text="Latest Announcements", text_color=COLORS["gold"], font=FONTS["h2"]).pack(anchor="w", padx=12, pady=6)
        cols2 = ("Message", "Date", "Sender")
        tv2 = ttk.Treeview(ann_frame, columns=cols2, show="headings", height=5)
        for c in cols2:
            tv2.heading(c, text=c)
            tv2.column(c, width=220, anchor="w")
        tv2.pack(fill="both", expand=True, padx=12, pady=8)
        style_treeview(tv2)
        for m, d, s in self.db.list_messages_for("all")[:10]:
            tv2.insert('', 'end', values=(m, d, s))
        self.bottom_frames.append(ann_frame)


class StudentsView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db
        top = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        top.pack(fill="x", padx=12, pady=8)
        self.search = ctk.CTkEntry(top, placeholder_text="Search by name/class/batch")
        self.search.pack(side="left", padx=6)
        ctk.CTkButton(top, text="Search", command=self.refresh, fg_color=COLORS["gold"], text_color=COLORS["bg1"]).pack(side="left", padx=6)
        GoldButton(top, text="Add Student", command=self._add_dialog).pack(side="right", padx=6)

        # table
        table_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        table_frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("ID", "Username", "Password", "Name", "Age", "Class", "Contact", "Email", "Batch", "Parent Contact", "Student Contact")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self.table.heading(c, text=c)
            self.table.column(c, width=120, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        style_treeview(self.table)

        # actions
        actions = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        actions.pack(fill="x", padx=12, pady=(0, 10))
        GoldButton(actions, text="Edit Selected", command=self._edit_selected).pack(side="left", padx=6)
        ctk.CTkButton(actions, text="Delete Selected", fg_color="#7a1f1f", hover_color="#953232", command=self._delete_selected).pack(side="left", padx=6)
        ctk.CTkButton(actions, text="Reset Password", fg_color="#444444", hover_color="#555555", command=self._reset_password).pack(side="left", padx=6)

    def _batch_names(self):
        return [name for name, _subj, _time in self.db.list_batches()]

    def refresh(self):
        for i in self.table.get_children():
            self.table.delete(i)
        rows = self.db.list_students_full_order(self.search.get().strip())
        for r in rows:
            self.table.insert('', 'end', values=r)

    def _add_dialog(self):
        Dialogs.student_form(self, title="Add Student", on_submit=self._add_student, batch_options=self._batch_names())

    def _add_student(self, data):
        # Validate mandatory parent contact
        if not data.get("parent_contact"):
            messagebox.showwarning("Validation", "Parent's phone number is required.")
            return
        try:
            self.db.add_student(data)
            messagebox.showinfo("Students", "Student added")
            self.refresh()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _edit_selected(self):
        item = self.table.focus()
        if not item:
            messagebox.showwarning("Edit", "Select a student row")
            return
        vals = self.table.item(item, 'values')
        sid = int(vals[0])
        data = {
            "username": vals[1],
            "password": vals[2],
            "name": vals[3],
            "age": int(vals[4] or 0),
            "class": vals[5],
            "contact": vals[6],
            "email": vals[7],
            "batch": vals[8],
            "parent_contact": vals[9],
            "student_contact": vals[10],
        }
        Dialogs.student_form(self, title="Edit Student", initial=data,
                             on_submit=lambda d: self._do_update(sid, d), batch_options=self._batch_names())

    def _do_update(self, sid, data):
        if not data.get("parent_contact"):
            messagebox.showwarning("Validation", "Parent's phone number is required.")
            return
        try:
            self.db.update_student(sid, data)
            messagebox.showinfo("Students", "Updated")
            self.refresh()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _delete_selected(self):
        item = self.table.focus()
        if not item:
            return
        if not messagebox.askyesno("Delete", "Delete selected student?"):
            return
        sid = int(self.table.item(item, 'values')[0])
        self.db.delete_student(sid)
        self.refresh()

    def _reset_password(self):
        item = self.table.focus()
        if not item:
            messagebox.showwarning("Reset", "Select a student row")
            return
        sid = int(self.table.item(item, 'values')[0])
        win = ctk.CTkToplevel(self)
        win.title("Reset Password"); win.geometry("360x160"); win.lift()
        ctk.CTkLabel(win, text="New Password:").pack(pady=(16,4))
        entry = ctk.CTkEntry(win, show='*'); entry.pack(padx=12, pady=4, fill='x')
        def do():
            pw = entry.get().strip()
            if not pw:
                return
            # minimal update: fetch existing row to keep other fields
            row = None
            for r in self.db.list_students_full_order():
                if r[0] == sid:
                    row = r; break
            if not row:
                win.destroy(); return
            data = {
                "username": row[1], "password": pw, "name": row[3], "age": row[4], "class": row[5],
                "contact": row[6], "email": row[7], "batch": row[8], "parent_contact": row[9], "student_contact": row[10]
            }
            self.db.update_student(sid, data)
            messagebox.showinfo("Reset", "Password reset")
            win.destroy()
        GoldButton(win, text="Save", command=do).pack(pady=12)


class BatchesView(ctk.CTkFrame):
    SUBJECT_CHOICES = ["Maths", "Science", "English", "Physics", "Chemistry", "Biology"]
    TIME_CHOICES = ["4-5", "6-7", "7-8"]

    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db

        top = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        top.pack(fill="x", padx=12, pady=8)

        # New batch form (Name, Subject dropdown, Time dropdown)
        form = ctk.CTkFrame(top, fg_color=COLORS["panel"], corner_radius=12)
        form.pack(side="left", padx=6, pady=6)
        self.name = ctk.CTkEntry(form, placeholder_text="Batch name", width=180)
        self.name.grid(row=0, column=1, padx=8, pady=6)
        ctk.CTkLabel(form, text="Name:", text_color=COLORS["gold"]).grid(row=0, column=0, sticky="e")
        ctk.CTkLabel(form, text="Subject:", text_color=COLORS["gold"]).grid(row=1, column=0, sticky="e")
        ctk.CTkLabel(form, text="Time:", text_color=COLORS["gold"]).grid(row=2, column=0, sticky="e")
        self.subj = ctk.CTkComboBox(form, values=self.SUBJECT_CHOICES, width=180)
        self.subj.grid(row=1, column=1, padx=8, pady=6)
        self.time = ctk.CTkComboBox(form, values=self.TIME_CHOICES, width=180)
        self.time.grid(row=2, column=1, padx=8, pady=6)
        GoldButton(form, text="Save Batch", command=self._save).grid(row=3, column=0, columnspan=2, pady=8)

        # Dropdown listing of existing batches
        right = ctk.CTkFrame(top, fg_color=COLORS["panel"], corner_radius=12)
        right.pack(side="left", padx=6, pady=6)
        ctk.CTkLabel(right, text="Batches:", text_color=COLORS["gold"]).pack(anchor="w", padx=8, pady=(8, 0))
        self.batch_dropdown = ttk.Combobox(right, state="readonly", width=40)
        self.batch_dropdown.pack(padx=8, pady=8)

        table_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        table_frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("Batch", "Subject", "Time")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self.table.heading(c, text=c)
            self.table.column(c, width=160, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        style_treeview(self.table)
        ctk.CTkButton(self, text="Delete Selected", fg_color="#7a1f1f", hover_color="#953232", command=self._delete).pack(pady=8)

    def refresh(self):
        # refresh table
        for i in self.table.get_children():
            self.table.delete(i)
        batches = self.db.list_batches()
        for name, subj, time in batches:
            self.table.insert('', 'end', values=(name, subj, time or "-"))
        # refresh dropdown
        items = [f"{name} — {subj} — {time or '-'}" for name, subj, time in batches]
        self.batch_dropdown["values"] = items
        if items:
            self.batch_dropdown.current(0)

    def _save(self):
        n = self.name.get().strip()
        s = (self.subj.get() or "").strip()
        t = (self.time.get() or "").strip()
        if not n:
            messagebox.showwarning("Validation", "Batch name is required")
            return
        self.db.upsert_batch(n, s, t)
        self.refresh()

    def _delete(self):
        item = self.table.focus()
        if not item:
            return
        name = self.table.item(item, 'values')[0]
        if messagebox.askyesno("Delete", f"Delete batch '{name}'?"):
            self.db.delete_batch(name)
            self.refresh()


class AttendanceView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db
        top = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        top.pack(fill="x", padx=12, pady=8)
        self.date = ctk.CTkEntry(top, placeholder_text="Date (YYYY-MM-DD)")
        self.date.pack(side="left", padx=6)
        self.status = ctk.CTkSegmentedButton(top, values=["Present", "Absent"])
        self.status.pack(side="left", padx=6)
        GoldButton(top, text="Mark Selected", command=self._mark).pack(side="left", padx=6)

        table_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        table_frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("ID", "Name", "Batch")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self.table.heading(c, text=c)
            self.table.column(c, width=140, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        style_treeview(self.table)

    def refresh(self):
        for i in self.table.get_children():
            self.table.delete(i)
        for r in self.db.list_students():
            self.table.insert('', 'end', values=(r[0], r[1], r[7]))

    def _mark(self):
        item = self.table.focus()
        if not item:
            return
        sid = int(self.table.item(item, 'values')[0])
        date = self.date.get().strip() or datetime.date.today().isoformat()
        status = self.status.get() or "Present"
        self.db.mark_attendance(sid, date, status)
        messagebox.showinfo("Attendance", "Saved")


class FeesView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db

        # Form
        form = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        form.pack(fill="x", padx=12, pady=8)

        # ID input and matches dropdown
        ctk.CTkLabel(form, text="Student ID:", text_color=COLORS["gold"]).grid(row=0, column=0, sticky="e", padx=6, pady=6)
        self.sid = ctk.CTkEntry(form, width=140)
        self.sid.grid(row=0, column=1, sticky="w", padx=6, pady=6)
        self.sid.bind("<KeyRelease>", self._on_id_change)
        self.matches = ttk.Combobox(form, state="readonly", width=40)
        self.matches.grid(row=0, column=2, padx=6, pady=6)
        self.matches.bind("<<ComboboxSelected>>", self._on_match_pick)

        # Read-only auto-filled fields
        self.name_var = ctk.StringVar(value="-")
        self.class_var = ctk.StringVar(value="-")
        self.batch_var = ctk.StringVar(value="-")
        ctk.CTkLabel(form, text="Name:", text_color=COLORS["gold"]).grid(row=1, column=0, sticky="e")
        ctk.CTkLabel(form, textvariable=self.name_var).grid(row=1, column=1, sticky="w")
        ctk.CTkLabel(form, text="Class:", text_color=COLORS["gold"]).grid(row=2, column=0, sticky="e")
        ctk.CTkLabel(form, textvariable=self.class_var).grid(row=2, column=1, sticky="w")
        ctk.CTkLabel(form, text="Batch:", text_color=COLORS["gold"]).grid(row=3, column=0, sticky="e")
        ctk.CTkLabel(form, textvariable=self.batch_var).grid(row=3, column=1, sticky="w")

        # Payment fields (do not auto-change)
        ctk.CTkLabel(form, text="Amount Paid:", text_color=COLORS["gold"]).grid(row=4, column=0, sticky="e")
        self.paid = ctk.CTkEntry(form, width=160)
        self.paid.grid(row=4, column=1, sticky="w", padx=6, pady=6)
        ctk.CTkLabel(form, text="Amount Pending:", text_color=COLORS["gold"]).grid(row=5, column=0, sticky="e")
        self.pending = ctk.CTkEntry(form, width=160)
        self.pending.grid(row=5, column=1, sticky="w", padx=6, pady=6)

        GoldButton(form, text="Record Payment", command=self._save).grid(row=6, column=0, columnspan=3, pady=10)

        self.total_label = ctk.CTkLabel(self, text="", font=FONTS["h2"], text_color=COLORS["gold"]) 
        self.total_label.pack(pady=8)

    def refresh(self):
        total = self.db.get_fees() or 0
        self.total_label.configure(text=f"Total Collected: ₹ {total:.2f}")

    def _on_id_change(self, _evt=None):
        prefix = self.sid.get().strip()
        if not prefix:
            self.matches["values"] = []
            self.name_var.set("-")
            self.class_var.set("-")
            self.batch_var.set("-")
            return
        matches = self.db.search_students_by_id_prefix(prefix)
        if not matches:
            self.matches["values"] = []
            return
        items = [f"{sid} — {name} ({username})" for sid, name, cls, username, batch in matches]
        self._matches_cache = {f"{sid} — {name} ({username})": sid for sid, name, cls, username, batch in matches}
        self.matches["values"] = items
        if len(items) == 1:
            self.matches.current(0)
            self._on_match_pick()

    def _on_match_pick(self, _evt=None):
        if not getattr(self, "_matches_cache", None):
            return
        label = self.matches.get()
        sid = self._matches_cache.get(label)
        if not sid:
            return
        rec = self.db.get_student_by_id(int(sid))
        if rec:
            # id, name, age, class, contact, email, username, batch, parent_contact, student_contact
            self.name_var.set(rec[1])
            self.class_var.set(rec[3])
            self.batch_var.set(rec[7])
            # leave paid/pending untouched

    def _save(self):
        try:
            sid = int(self.sid.get())
            paid = float(self.paid.get() or 0)
            pend = float(self.pending.get() or 0)
            date = datetime.date.today().isoformat()
            self.db.record_payment(sid, paid, pend, date)
            messagebox.showinfo("Fees", "Payment recorded")
            self.refresh()
        except Exception as e:
            messagebox.showerror("Fees", str(e))


class TimetableView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db

        # Controls
        top = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        top.pack(fill="x", padx=12, pady=8)
        ctk.CTkLabel(top, text="Batch:", text_color=COLORS["gold"]).pack(side="left")
        self.batch = ctk.CTkComboBox(top, values=[b for b,_s,_t in self.db.list_batches()] or [""], width=160)
        if self.db.list_batches():
            self.batch.set(self.db.list_batches()[0][0])
        self.batch.pack(side="left", padx=6)
        GoldButton(top, text="Import Excel", command=lambda: self._import("xlsx")).pack(side="left", padx=6)
        GoldButton(top, text="Import CSV", command=lambda: self._import("csv")).pack(side="left", padx=6)
        ctk.CTkButton(top, text="Clear Batch", fg_color="#7a1f1f", hover_color="#953232", command=self._clear_batch).pack(side="left", padx=6)
        GoldButton(top, text="Export Excel", command=lambda: self._export("xlsx")).pack(side="left", padx=6)
        GoldButton(top, text="Export PDF", command=lambda: self._export("pdf")).pack(side="left", padx=6)

        # Table
        table_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        table_frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("ID","Batch","Day","Time","Subject","TeacherID")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self.table.heading(c, text=c)
            self.table.column(c, width=120, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        style_treeview(self.table)
        ctk.CTkButton(self, text="Delete Selected", fg_color="#7a1f1f", hover_color="#953232", command=self._delete).pack(pady=8)

    def refresh(self):
        for i in self.table.get_children():
            self.table.delete(i)
        batch = self.batch.get().strip() if hasattr(self,'batch') else None
        rows = self.db.list_timetable(batch if batch else None)
        for r in rows:
            self.table.insert('', 'end', values=r)

    def _clear_batch(self):
        b = self.batch.get().strip()
        if not b:
            return
        from tkinter import messagebox
        if not messagebox.askyesno("Clear", f"Clear timetable for batch '{b}'?"):
            return
        self.db.clear_timetable_for_batch(b)
        self.refresh()

    def _import(self, fmt: str):
        from tkinter import filedialog
        try:
            b = self.batch.get().strip()
            if not b:
                return
            path = None
            if fmt == "xlsx":
                path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
            else:
                path = filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
            if not path:
                return
            # Clear current batch entries
            self.db.clear_timetable_for_batch(b)
            if fmt == "xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(path); ws = wb.active
                # Expect header row with Batch,Day,Time,Subject,TeacherID
                first = True
                for row in ws.iter_rows(values_only=True):
                    if first:
                        first = False; continue
                    batch, day, tm, subj, tid = row[:5]
                    if not batch: batch = b
                    tidv = int(tid) if tid not in (None, "") else None
                    self.db.upsert_timetable_entry(str(batch), str(day), str(tm), str(subj), tidv)
            else:
                import csv as _csv
                with open(path, newline='', encoding='utf-8') as f:
                    r = _csv.reader(f)
                    first = True
                    for row in r:
                        if first:
                            first = False; continue
                        batch, day, tm, subj, tid = (row + [None]*5)[:5]
                        if not batch: batch = b
                        tidv = int(tid) if tid and tid.strip().isdigit() else None
                        self.db.upsert_timetable_entry(batch, day, tm, subj, tidv)
            from tkinter import messagebox
            messagebox.showinfo("Import", "Timetable imported")
            self.refresh()
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Import", str(e))

    def _export(self, format: str = "xlsx"):
        try:
            rows = self.db.list_timetable(self.batch.get().strip())
            if format == "xlsx":
                try:
                    from openpyxl import Workbook
                    from tkinter import filedialog
                    path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel",".xlsx")])
                    if not path:
                        return
                    wb = Workbook(); ws = wb.active; ws.append(["ID","Batch","Day","Time","Subject","TeacherID"])
                    for r in rows: ws.append(list(r))
                    wb.save(path)
                    messagebox.showinfo("Export","Excel saved")
                except Exception as e:
                    messagebox.showerror("Export", f"Excel export failed: {e}")
            else:
                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.pdfgen import canvas
                    from tkinter import filedialog
                    path = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF",".pdf")])
                    if not path:
                        return
                    c = canvas.Canvas(path, pagesize=A4)
                    x, y = 40, 800
                    c.setFont("Helvetica-Bold", 14); c.drawString(x, y, f"Timetable - {self.batch.get().strip()}"); y -= 24
                    c.setFont("Helvetica", 11)
                    for r in rows:
                        c.drawString(x, y, f"{r[0]}  {r[1]}  {r[2]}  {r[3]}  {r[4]}  {r[5] or ''}")
                        y -= 16
                        if y < 60:
                            c.showPage(); y = 800
                    c.save()
                    messagebox.showinfo("Export","PDF saved")
                except Exception as e:
                    messagebox.showerror("Export", f"PDF export failed: {e}")
        except Exception as e:
            messagebox.showerror("Export", str(e))

    def _delete(self):
        item = self.table.focus()
        if not item:
            return
        vid = int(self.table.item(item, 'values')[0])
        self.db.delete_timetable_entry(vid)
        self.refresh()


class MessagesView(ctk.CTkFrame):
    def __init__(self, master, db):
        super().__init__(master, fg_color=COLORS["bg1"]) 
        self.db = db
        top = ctk.CTkFrame(self, fg_color=COLORS["bg1"]) 
        top.pack(fill="x", padx=12, pady=8)
        self.message = ctk.CTkEntry(top, placeholder_text="Type message (to all)", width=420)
        self.message.pack(side="left", padx=6)
        GoldButton(top, text="Send", command=self._send).pack(side="left", padx=6)

        # History table (all messages)
        table_frame = ctk.CTkFrame(self, fg_color=COLORS["panel"], corner_radius=12)
        table_frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("Message", "Date", "Sender", "Recipient")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings")
        for c in cols:
            self.table.heading(c, text=c)
            self.table.column(c, width=220, anchor="w")
        self.table.pack(fill="both", expand=True, padx=8, pady=8)
        style_treeview(self.table)

    def refresh(self):
        for i in self.table.get_children():
            self.table.delete(i)
        for m, d, s, r in self.db.list_all_messages():
            self.table.insert('', 'end', values=(m, d, s, r))

    def _send(self):
        text = self.message.get().strip()
        if not text:
            return
        self.db.send_message(text, datetime.datetime.now().isoformat(sep=' ', timespec='seconds'), "admin", "all")
        self.message.delete(0, 'end')
        messagebox.showinfo("Messages", "Sent")


class Dialogs:
    @staticmethod
    def _labeled_entry(parent, label: str, initial: str = "", password: bool = False):
        row = ctk.CTkFrame(parent, fg_color=COLORS["panel"])
        row.pack(fill="x", padx=12, pady=4)
        ctk.CTkLabel(row, text=label, width=120, anchor="w", text_color=COLORS["gold"]).pack(side="left")
        entry = ctk.CTkEntry(row, show="*" if password else None)
        entry.pack(side="left", fill="x", expand=True, padx=8)
        if initial:
            entry.insert(0, initial)
        return entry

    @staticmethod
    def student_form(parent, title: str, on_submit, initial=None, batch_options=None):
        initial = initial or {}
        batch_options = batch_options or []
        win = ctk.CTkToplevel(parent)
        win.title(title)
        win.geometry("520x600")
        # Ensure on top and modal
        win.lift()
        try:
            win.attributes("-topmost", True)
        except Exception:
            pass
        try:
            win.transient(parent.winfo_toplevel())
            win.grab_set()
        except Exception:
            pass

        fields = [
            ("name", "Name:"),
            ("age", "Age:"),
            ("class", "Class:"),
            ("contact", "Contact:"),
            ("email", "Email:"),
            ("username", "Username:"),
            ("password", "Password:", True),
            ("parent_contact", "Parent Phone (required):"),
            ("student_contact", "Student Phone (optional):"),
        ]
        for key, label, *rest in fields:
            pwd = bool(rest and rest[0])
            val = initial.get(key)
            if val is None:
                val = ""
            setattr(win, key, Dialogs._labeled_entry(win, label, str(val), password=pwd))

        # Batch combobox (labeled)
        row = ctk.CTkFrame(win, fg_color=COLORS["panel"]) 
        row.pack(fill="x", padx=12, pady=4)
        ctk.CTkLabel(row, text="Batch:", width=120, anchor="w", text_color=COLORS["gold"]).pack(side="left")
        win.batch = ctk.CTkComboBox(row, values=batch_options, width=240)
        if initial.get("batch"):
            try:
                win.batch.set(initial.get("batch"))
            except Exception:
                pass
        win.batch.pack(side="left", padx=8)

        def submit():
            data = {k: getattr(win, k).get().strip() for k, _l, *r in fields}
            data["batch"] = win.batch.get().strip() if hasattr(win, "batch") else initial.get("batch", "")
            try:
                data["age"] = int(data.get("age") or 0)
            except Exception:
                data["age"] = 0
            on_submit(data)
            win.destroy()

        GoldButton(win, text="Save", command=submit).pack(pady=12)
